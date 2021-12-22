from tkinter import *
from tkinter import messagebox
from openpyxl import *
import time
import random
from imageRecognizer import imageRecognizer
from rfidReader import rfidRead


#========================Tunable Params===========================
recogLoopCount=3
recogConfidence=95


# =========================BACK-END================================


def error_handling(msg):
    if msg == "error!":
        messagebox.showerror("Error", "No Digits to compute!")
    elif msg == "division by zero!":
        messagebox.showerror("Error", "You can not divide a digit by zero!")



def exit_program():
    root.destroy()


def reset_sales():
    try:
        check_btn_var1.set(0)
        check_btn_var2.set(0)
        check_btn_var3.set(0)
        check_btn_var4.set(0)
        check_btn_var5.set(0)
        check_btn_var6.set(0)
        check_btn_var7.set(0)
        check_btn_var8.set(0)
        menu_opt_var1.set(list_menu[0])
        menu_opt_var2.set(list_menu[0])
        menu_opt_var3.set(list_menu[0])
        menu_opt_var4.set(list_menu[0])
        menu_opt_var5.set(list_menu[0])
        menu_opt_var6.set(list_menu[0])
        menu_opt_var7.set(list_menu[0])
        menu_opt_var8.set(list_menu[0])
        


        entry_total_cost.delete(0, END)
       
        
        receipt_field.delete("1.0", END)
        field.delete("1.0", END)
        scanneditems.clear()
        messagebox.showinfo("Next Customer!", "Serve the next customer please!")
    except NameError as e:
        print(e)
        messagebox.showerror("Error!", "Nothing to clear!")


def total_sales():
    # ================================ FRUITS ============================== #
    price_tag = {"Casio": 2500, "Chocolatto box": 120, "Lays French Cheese": 30, "UV Matt": 1000, "HIK VISION SSD": 3650, "Colgate": 350, "Khopra Candy Packet": 120,
                 "Shan Chicken 65": 75,"other": entry_price_other_fruit_var.get()}

    global items_quantity_price
    items_quantity_price = []

    if  check_btn_var1.get() == 1:
        casio_results = int(menu_opt_var1.get()) * price_tag["Casio"]
        casio_tuple = ("Casio", str(menu_opt_var1.get()), "2500")
        items_quantity_price.append(casio_tuple)
    else:
        casio_results = 0

    if check_btn_var2.get() == 1:
        choco_results = int(menu_opt_var2.get()) * price_tag["Chocolatto box"]
        choco_tuple = ("Chocolatto box", str(menu_opt_var2.get()), "120")
        items_quantity_price.append(choco_tuple)
    else:
        choco_results = 0

    if check_btn_var3.get() == 1:
        lays_results = int(menu_opt_var3.get()) * price_tag["Lays French Cheese"]
        lays_tuple = ("Lays French Cheese", str(menu_opt_var3.get()), "25")
        items_quantity_price.append(lays_tuple)
    else:
        lays_results = 0

    if check_btn_var4.get() == 1:
        uv_results = int(menu_opt_var4.get()) * price_tag["UV Matt"]
        uv_tuple = ("UV Matt", str(menu_opt_var4.get()), "1000")
        items_quantity_price.append(uv_tuple)
    else:
        uv_results = 0

    if check_btn_var5.get() == 1:
        hv_results = int(menu_opt_var5.get()) * price_tag["HIK VISION SSD"]
        hv_tuple = ("HIK VISION SSD", str(menu_opt_var5.get()), "3650")
        items_quantity_price.append(hv_tuple)
    else:
        hv_results = 0

    if check_btn_var6.get() == 1:
        Colgate_results = int(menu_opt_var6.get()) * price_tag["Colgate"]
        Colgate_tuple = ("Colgate", str(menu_opt_var6.get()), "350")
        items_quantity_price.append(Colgate_tuple)
    else:
        Colgate_results = 0

    if check_btn_var7.get() == 1:
        kc_results = int(menu_opt_var7.get()) * price_tag["Khopra Candy Packet"]
        kc_tuple = ("Khopra Candy Packet", str(menu_opt_var7.get()), "120")
        items_quantity_price.append(kc_tuple)
    else:
        kc_results = 0

    if check_btn_var8.get() == 1:
        sc_results = int(menu_opt_var8.get()) * price_tag["Shan Chicken 65"]
        sc_tuple = ("Shan Chicken 65", str(menu_opt_var8.get()), "75")
        items_quantity_price.append(sc_tuple)
    else:
        sc_results = 0


    try:
        global other_item_results
        other_item_results=0
        global total_cost
        

        total_cost= (casio_results + choco_results + lays_results + uv_results + hv_results +
                                Colgate_results + kc_results + sc_results + other_item_results  )

        if total_cost>0:
            report_btn_btm_frame['state']=NORMAL

        entry_total_cost.delete(0, END)
        entry_total_cost.insert(0, "KES. " + str(total_cost))
    except NameError as e:
        print(e)
        error_handling("error!")
    except ValueError:
        messagebox.showerror("Error!", "Price your items first, before finding the total.")

        
        subtotal_entry.delete(0, END)
        subtotal_entry.insert(0, "KES. " + str(total_cost_of_juice + total_cost_of_fruits))

        total_entry.delete(0, END)
        total_entry.insert(0, "KES. " + str(total_cost_of_juice + total_cost_of_fruits + tax_on_all_items))
    except NameError as e:
        print(e)
        error_handling("error!")
    except ValueError:
        messagebox.showerror("Error!", "Price your items first, before finding the total.")


def print_receipt():
    total_sales()
    try:
        receipt_field.delete("1.0", END)
        receipt_field.insert(END, "  " + "\n")  # blank space
        receipt_field.insert(END, "  " + "\n")  # blank space
        receipt_field.insert(END, "  Autonomous Trolley With Smart Billing." + "\n")
        receipt_field.insert(END, " " + "\n")

        x = 0
        while True:
            receipt_field.insert(END, items_quantity_price[x][0] + " :   " + items_quantity_price[x][1] + " each @ " +
                                 items_quantity_price[x][2] + "\n")
            x += 1

            if x < len(items_quantity_price):
                continue
            else:
                break

        receipt_field.insert(END, " " + "\n")
        receipt_field.insert(END, "Total cost:  " + entry_total_cost_var.get() + "\n")
        receipt_field.insert(END, " " + "\n")
        receipt_field.insert(END, " ***** WELCOME AGAIN *****" + "\n")
    except IndexError as e:
        print("Index error "+ e)
        messagebox.showerror("Error!", "select items first! Thank you!")
    except NameError as e:
        print("Name error " + e)
        messagebox.showerror("Error!", "select items first! Thank you!")


def book_keeping():
    # LOAD THE WORKBOOK
    try:
        global book
        destination_file = "SmartTrolley.xlsx"
        book = load_workbook(filename=destination_file)

        # GRABBING THE ACTIVE SHEET.
        sheet = book.active

        # USEFUL VARIABLES.
        row = 2
        column = 1
        global today
        today = time.strftime("%x")
        id_ = random.randint(1, 1000000)

        if items_quantity_price:
            while True:
                if sheet.cell(row=row, column=column).value:
                    row += 1
                    continue
                else:
                    sheet.cell(row=row, column=column).value = today
                    column += 1
                    sheet.cell(row=row, column=column).value = id_
                    column += 1

                    # adding elements to the item cell
                    x = len(items_quantity_price)
                    sheet.cell(row=row, column=column).value = x
                    column += 1

                    sheet.cell(row=row, column=column).value = total_cost_of_juice + total_cost_of_fruits
                    column += 1
                    sheet.cell(row=row, column=column).value = tax_on_all_items

                    if column == 5:
                        break

            book.save(filename=destination_file)
            messagebox.showinfo("Book-keeping", "DONE!")
        else:
            messagebox.showerror("Error!", "Nothing to add!")
    except NameError:
        messagebox.showerror("Error!", "Nothing to add!")


    
def camera_scanning():
    global scannedIndex
    global scanneditems
    scannedIndex=scannedIndex+1

    print(scannedIndex)
    predictedclass= imageRecognizer(minPredictions=recogLoopCount, averageConfidence=recogConfidence)
    
    
    if predictedclass== '0':
        i=int(menu_opt_var1.get())
        menu_opt_var1.set(list_menu[i+1])
        check_btn1.select()
        scanneditems[str(scannedIndex)]="Casio Rs:2500"

    if predictedclass=='1':
        i=int(menu_opt_var2.get())
        menu_opt_var2.set(list_menu[i+1])
        check_btn2.select()
        scanneditems[str(scannedIndex)]="Chocolatto Box Rs:120"
    
    if predictedclass=='5':
        i=int(menu_opt_var3.get())
        menu_opt_var3.set(list_menu[i+1])
        check_btn3.select()
        scanneditems[str(scannedIndex)]="Lays 30 Rs:30"
    
    if predictedclass=='7':
        i=int(menu_opt_var4.get())
        menu_opt_var4.set(list_menu[i+1])
        check_btn4.select()
        scanneditems[str(scannedIndex)]="UV Sunblock Rs:1000"

    if predictedclass=='3':
        i=int(menu_opt_var5.get())
        menu_opt_var5.set(list_menu[i+1])
        check_btn5.select()
        scanneditems[str(scannedIndex)]="Hikvision SSD Rs:3650"

    if predictedclass=='2':
        i=int(menu_opt_var6.get())
        menu_opt_var6.set(list_menu[i+1])
        check_btn6.select()
        scanneditems[str(scannedIndex)]="Colgate Tootpaste Rs:350"

    if predictedclass=='4':
        i=int(menu_opt_var7.get())
        menu_opt_var7.set(list_menu[i+1])
        check_btn7.select()
        scanneditems[str(scannedIndex)]="Khopra Candy Packet Rs:120"

    if predictedclass=='6':
        i=int(menu_opt_var8.get())
        menu_opt_var8.set(list_menu[i+1])
        check_btn8.select()
        scanneditems[str(scannedIndex)]="Shan Chicken 65 Rs:75"


    field.delete("1.0", END)
    field.insert(END, "  " + "\n")  # blank space
    field.insert(END, "  " + "\n")  # blank space
    field.insert(END, " You can select/deselect items from left bar " + "\n")
    field.insert(END, " " + "\n")
    scannedString=""
    for index in scanneditems:
        scannedString+=str(index)
        scannedString+= " : "
        scannedString+= scanneditems[index]
        scannedString+= "\n"
    
    field.insert(END, scannedString)
    total_sales()
    
    
def rfid_read():
    predictedclass= rfidRead()
    
    
    global scannedIndex
    global scanneditems
    scannedIndex=scannedIndex+1

    print(scannedIndex) 
    if predictedclass== '1':
        i=int(menu_opt_var1.get())
        menu_opt_var1.set(list_menu[i+1])
        check_btn1.select()
        scanneditems[str(scannedIndex)]="Casio Rs:2500"

    if predictedclass=='2':
        i=int(menu_opt_var2.get())
        menu_opt_var2.set(list_menu[i+1])
        check_btn2.select()
        scanneditems[str(scannedIndex)]="Chocolatto Box Rs:120"
    
    if predictedclass=='3':
        i=int(menu_opt_var3.get())
        menu_opt_var3.set(list_menu[i+1])
        check_btn3.select()
        scanneditems[str(scannedIndex)]="Lays 30 Rs:30"
    
    if predictedclass=='4':
        i=int(menu_opt_var4.get())
        menu_opt_var4.set(list_menu[i+1])
        check_btn4.select()
        scanneditems[str(scannedIndex)]="UV Sunblock Rs:1000"

    if predictedclass=='5':
        i=int(menu_opt_var5.get())
        menu_opt_var5.set(list_menu[i+1])
        check_btn5.select()
        scanneditems[str(scannedIndex)]="Hikvision SSD Rs:3650"

    if predictedclass=='6':
        i=int(menu_opt_var6.get())
        menu_opt_var6.set(list_menu[i+1])
        check_btn6.select()
        scanneditems[str(scannedIndex)]="Colgate Tootpaste Rs:350"

    if predictedclass=='7':
        i=int(menu_opt_var7.get())
        menu_opt_var7.set(list_menu[i+1])
        check_btn7.select()
        scanneditems[str(scannedIndex)]="Khopra Candy Packet Rs:120"

    if predictedclass=='8':
        i=int(menu_opt_var8.get())
        menu_opt_var8.set(list_menu[i+1])
        check_btn8.select()
        scanneditems[str(scannedIndex)]="Shan Chicken 65 Rs:75"

    if predictedclass== '9':
        messagebox.showerror("Error","Error Reading the tag")
    if predictedclass=='10':
        messagebox.showerror("Error","Invalid ID: product does not exist")
        


    field.delete("1.0", END)
    field.insert(END, "  " + "\n")  # blank space
    field.insert(END, "  " + "\n")  # blank space
    field.insert(END, " You can select/deselect items from left bar " + "\n")
    field.insert(END, " " + "\n")
    scannedString=""
    for index in scanneditems:
        scannedString+=str(index)
        scannedString+= " : "
        scannedString+= scanneditems[index]
        scannedString+= "\n"
    
    field.insert(END, scannedString)
    total_sales()

def today_report():
    destination_file2 = "POS_fruit_juice.xlsx"
    work_book = load_workbook(filename=destination_file2)

    # GRABBING THE ACTIVE SHEET.
    sheet2 = work_book.active

    row = 2
    column = 1
    max_row = sheet2.max_row
    total_amount = []
    total_vat = []
    today2 = time.strftime("%x")

    x = 0
    while x < max_row:
        if sheet2.cell(row=row, column=column).value == today2:
            total_amount.append(sheet2.cell(row=row, column=4).value)
            total_vat.append(sheet2.cell(row=row, column=5).value)
            row += 1
            x += 1
            continue
        else:
            row += 1
            x += 1
            continue

    answer1 = 0
    for integer in total_amount:
        answer1 += integer

    answer2 = 0
    for float_point in total_vat:
        answer2 += float_point

    receipt_field.delete("1.0", END)
    receipt_field.insert(END, "Autonomous Trolley With Smart Billing.          " + today2 + "." + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space
    receipt_field.insert(END, "         ***TODAY\'S REPORT.***" + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space
    receipt_field.insert(END, "Total Amount generated today: " + "KES. " + str(answer1) + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space
    receipt_field.insert(END, "Total V.A.T generated today: " + "KES. " + str(answer2) + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space
    receipt_field.insert(END, "Gross income + V.A.T: " + "KES. " + str(answer1 + answer2) + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space

# ======================BACK-END=============================

#======================Specific Product Deletion==========    

            
def product_deletion():
    global scannedIndex
    global scanneditems



    
    deletion=Toplevel(root)
    deletion.title("DeleteProducts")
    deletion.resizable(width=FALSE, height=FALSE)
    deletion.geometry('580x400')

    def deleteSelected(checkArray):
        itemDeletion= [None]* len(checkArray)
        for i, item in enumerate(scanneditems):
            if checkArray[i].get()==1:
                itemDeletion[i]=item
        print(itemDeletion)
        for i in itemDeletion:
            if i!=None:
                deletedItem= scanneditems[i]
                if deletedItem== 'Casio Rs:2500':
                    j=int(menu_opt_var1.get())
                    menu_opt_var1.set(list_menu[j-1])

                if deletedItem=='Chocolatto Box Rs:120':
                    j=int(menu_opt_var2.get())
                    menu_opt_var2.set(list_menu[j-1])
                    
                
                if deletedItem=='Lays 30 Rs:30':
                    j=int(menu_opt_var3.get())
                    menu_opt_var3.set(list_menu[j-1])
                                    
                if deletedItem=='UV Sunblock Rs:1000':
                    j=int(menu_opt_var4.get())
                    menu_opt_var4.set(list_menu[j-1])

                if deletedItem=='Hikvision SSD Rs:3650':
                    i=int(menu_opt_var5.get())
                    menu_opt_var5.set(list_menu[j-1])
                    
                if deletedItem=='Colgate Tootpaste Rs:350':
                    j=int(menu_opt_var6.get())
                    menu_opt_var6.set(list_menu[j-1])

                if deletedItem=='Khopra Candy Packet Rs:120':
                    j=int(menu_opt_var7.get())
                    menu_opt_var7.set(list_menu[j-1])

                if deletedItem=='Shan Chicken 65 Rs:75':
                    j=int(menu_opt_var8.get())
                    menu_opt_var8.set(list_menu[j-1])

                del scanneditems[i]
                total_sales()
                

        field.delete("1.0", END)
        field.insert(END, "  " + "\n")  # blank space
        field.insert(END, "  " + "\n")  # blank space
        field.insert(END, " You can select/deselect items from left bar " + "\n")
        field.insert(END, " " + "\n")
        scannedString=""
        for index in scanneditems:
            scannedString+=str(index)
            scannedString+= " : "
            scannedString+= scanneditems[index]
            scannedString+= "\n"
        
        field.insert(END, scannedString)
        deletion.destroy()
        deletion.update()
    
    lft_frame = Frame(deletion, bg=color3, bd=4, relief=RIDGE, width=500, height=400)
    lft_frame.place(x=40, y=0)

    label_show =Label(deletion, text="Product Deletion",bd=7, relief=RIDGE,width=25, height=2,font=("Times", 15, "bold","italic"),bg=color4, fg=color)
    label_show.place(x=120,y=0)

    checkArray= [None] * len(scanneditems)
    for i,item in enumerate(scanneditems):
        checkArray[i] = IntVar()
        newButton = Checkbutton(lft_frame, text=scanneditems[item], bg=color4,bd=5,relief=RAISED , fg=color, font=("Candra", 10, "bold"),variable=checkArray[i])
        newButton.place(x=100, y=55 +(i*45))

    deletionButton = Button(lft_frame, text="Delete Selected", width=22, height=3,bg=color4, bd=7, relief=RAISED, font=("Arial", 12, "bold"), fg=color, command=lambda: deleteSelected(checkArray) ).place(x=150,y=300)



    deletion.configure(bg=color3)


#=======================Checkout==========================
def rfidDiscount(amount_show):
    global total_cost
    
    cardRead=rfidRead()

    if cardRead=="2%":
        total_cost= total_cost- (total_cost*2/100)
    elif cardRead=="5%":
        total_cost= total_cost- (total_cost*5/100)
    else:
        messagebox.showerror("Invalid Card" ," Error: invalid card please retry. ")

    amount_show["text"] = "Total: "+str(total_cost)
def checkout():
    total_sales()
    global total_cost
    checkoutWin= Tk()
    checkoutWin.title("Checking Out")
    checkoutWin.resizable(width=FALSE, height=FALSE)
    checkoutWin.geometry('580x400')

    lft_frame = Frame(checkoutWin, bg=color3, bd=4, relief=RIDGE, width=500, height=400)
    lft_frame.place(x=40, y=0)

    amount_show =Label(checkoutWin, text="Total: "+str(total_cost),bd=7, relief=RIDGE,width=25, height=2,font=("Times", 15, "bold","italic"),bg=color4, fg=color)
    
    label_show =Label(checkoutWin, text="Checking Out",bd=7, relief=RIDGE,width=25, height=2,font=("Times", 15, "bold","italic"),bg=color4, fg=color).place(x=120,y=30)
    voucherScan = Button(checkoutWin, text="Voucher Card Scanning", width=22, height=2,bg=color4, bd=7, relief=RAISED, font=("Arial", 9, "bold"), fg=color, command=lambda: rfidDiscount(amount_show) ).place(x=170,y=110)

    amount_show.place(x=120,y=170)
    paymentProceed = Button(checkoutWin, text="Proceed to Payment", width=30,height=2, bg=color4,bd=7,relief=RAISED,font=("Candra", 7,"bold") ,fg=color, command=lambda: second_win()).place(x=170,y=280)

    checkoutWin.configure(bg=color3)


#===========Second Window Frames==========================
def third_win():
    global total_cost
    window2=Tk()
    window2.title("Payment Method")
    window2.resizable(width=FALSE, height=FALSE)
    window2.geometry('580x400')

    # new window buttons
    lft_frame = Frame(window2, bg=color3, bd=4, relief=RIDGE, width=500, height=400)
    lft_frame.place(x=30, y=85)
    # receipt
    r_frame2 = Frame(window2, bg=color4, bd=3, relief=RIDGE)
    r_frame2.place(x=730, y=190)
    # ==============================Text field on right_frame2========================


    label_show =Label(window2, text="Please Proceed To Cash Counter",bd=7, relief=RIDGE,width=25, height=2,font=("Times", 20, "bold","italic"),bg=color4, fg=color).place(x=70,y=30)
    ll_show = Label(window2, text="Your Total Bill is "+ str(total_cost), bd=7, relief="solid", width=20, height=1,
                       font=("Times", 20, "bold", "italic"), bg=color4, fg=color).place(x=100,y=140)
    window2.configure(bg=color3)


#===========Payment Method Panel==========================
def second_win():
    global total_cost
    window=Tk()
    window.title("Payment Method")
    window.resizable(width=FALSE, height=FALSE)
    window.geometry('580x400')

    # new window buttons
    lft_frame = Frame(window, bg=color3, bd=4, relief=RIDGE, width=500, height=370)
    lft_frame.place(x=30, y=85)


    label_show =Label(window, text="Choose any Payment Method",bd=7, relief=RIDGE,width=25, height=2,font=("Times", 20, "bold","italic"),bg=color4, fg=color).place(x=70,y=30)
    click_1 = Button(lft_frame, text="Cash", width=15, height=5,bg=color4, bd=7, relief=RAISED, font=("Arial", 10, "bold"), fg=color, command=lambda: third_win()).place(x=60,y=70)
    click_2 = Button(lft_frame, text="Online Payment", width=15, height=5,bg=color4, bd=7, relief=RAISED, font=("Arial", 10, "bold"),
                     fg=color).place(x=280,y=70)

    click_3 = Button(lft_frame, text="Easiy Paisa", width=15, height=5, bg=color4, bd=7, relief=RAISED,
                     font=("Arial", 10, "bold"),
                     fg=color).place(x=60, y=230)

    click_4 = Button(lft_frame, text="Jazz Cash", width=15, height=5, bg=color4, bd=7, relief=RAISED,
                     font=("Arial", 10, "bold"),
                     fg=color).place(x=280, y=230)

    window.configure(bg=color3)



# ==========================Main Panel======================

root=Tk()
root.resizable(width=FALSE, height=FALSE)
root.geometry("800x600")
root.title("Autonomous Trolley")
root.attributes('-fullscreen', True)
color = "Silver"
color2 = "#0a2f49"
color3 = "#031c2d"
color4 = '#11517c'
root.configure(bg=color3)
total_cost=0
##scanned Items settings
scanneditems= dict()
scannedIndex=0

deletionIndex = StringVar()
deletionIndex.set("Select an Item")

# ==========================FRAMES=========================

# fruits
left_frame = Frame(root, bg=color3,bd=4,relief=RIDGE, width=800, height=600)
left_frame.place(x=0, y=0)

left_framefruit = Frame(root, bg=color2,bd=4,relief=RIDGE, width=250, height=400)
left_framefruit.place(x=10, y=40)


# billing results
right_frame = Frame(root, bg=color2, width=200, height=27)
right_frame.place(x=550,y=40)

# receipt
right_frame2 = Frame(root, bg=color4,bd=3,relief=RIDGE)
right_frame2.place(x=525, y=75)

#scanner
left_frame2 = Frame(root, bg=color2,bd=3,relief=RIDGE)
left_frame2.place(x=270, y=55)


# ==============================LABELS & ENTRIES on right frame========================

label_total_cost = Label(right_frame, text="Total Cost : ", font=("Candra", 10, "bold"), bg=color4,fg=color)
label_total_cost.place(x=0, y=0)

entry_total_cost_var = StringVar()
entry_total_cost = Entry(right_frame, textvariable=entry_total_cost_var, font=("Candra", 10, "bold"),bg=color2,bd=3,relief=RIDGE , fg=color, width=15)
entry_total_cost.place(x=80, y=0)

# ==============================Text field on right_frame2========================
receipt_field = Text(right_frame2, width=35, height=20, fg=color, font=("Candra", 8, "bold"), bg=color2)
receipt_field.pack(side=TOP)
label_recipt = Label(root, text=" Total Bill ", bg=color3,bd=3,relief=RIDGE, font=("Times", 13, "bold"), fg=color)
label_recipt.place(x=550, y=40)
# ==============================scanner bar on left_frame2========================
field = Text(left_frame2, width=40, height=23, fg=color, font=("Candra", 7, "bold"), bg=color2)
field.pack(side=TOP)

label_scan = Label(root, text=" Scan Products ", bg=color3,bd=3,relief=RIDGE, font=("Times", 13, "bold"), fg=color)
label_scan.place(x=310, y=50)
# ==============================Check-boxes, label, & an entry on the left_frame=====================


check_btn_var1 = IntVar()
check_btn1 = Checkbutton(left_framefruit, text="Casio", bg=color4,bd=5,relief=RAISED , fg=color, font=("Candra", 10, "bold"),variable=check_btn_var1)
check_btn1.place(x=7, y=20)


check_btn_var2 = IntVar()
check_btn2 = Checkbutton(left_framefruit, text="Chocolatto Box", bg=color4 ,bd=5,relief=RAISED, fg=color, font=("Candra", 10, "bold"), variable=check_btn_var2)
check_btn2.place(x=7, y=60)

check_btn_var3 = IntVar()
check_btn3 = Checkbutton(left_framefruit, text="Lays Fr Cheese", bg=color4,bd=5,relief=RAISED , fg=color, font=("Candra",10, "bold"), variable=check_btn_var3)
check_btn3.place(x=7, y=100)

check_btn_var4 = IntVar()
check_btn4 = Checkbutton(left_framefruit, text="UV Matt", bg=color4,bd=5,relief=RAISED , fg=color, font=("Candra", 10, "bold"), variable=check_btn_var4)
check_btn4.place(x=7, y=140)

check_btn_var5 = IntVar()
check_btn5 = Checkbutton(left_framefruit, text="HIK VISION SSD", bg=color4,bd=5,relief=RAISED , fg=color, font=("Candra",10, "bold"),variable=check_btn_var5)
check_btn5.place(x=7, y=180)

check_btn_var6 = IntVar()
check_btn6 = Checkbutton(left_framefruit, text="Colgate",bg=color4,bd=5,relief=RAISED , fg=color, font=("Candra", 10, "bold"), variable=check_btn_var6)
check_btn6.place(x=7, y=220)

check_btn_var7 = IntVar()
check_btn7 = Checkbutton(left_framefruit, text="Khopra Packet",bg=color4,bd=5,relief=RAISED , fg=color, font=("Candra", 10, "bold"), variable=check_btn_var7)
check_btn7.place(x=7, y=260)

check_btn_var8 = IntVar()
check_btn8 = Checkbutton(left_framefruit, text="Shan Chicken 65", bg=color4,bd=5,relief=RAISED , fg=color, font=("Candra", 10, "bold"),variable=check_btn_var8)
check_btn8.place(x=7, y=300)




# ======================menu-options on the left frame=============================
list_menu = ["0", "1", "2", "3", "4", "5", "6", "7", "8"]


menu_opt_var1 = StringVar()
menu_opt_var1.set(list_menu[0])
menu_opt1 = OptionMenu(left_framefruit, menu_opt_var1, *list_menu)
menu_opt1.configure(bg=color4,bd=1,relief=RIDGE , fg=color, font=("Times", 10))
menu_opt1.place(x=180, y=20)

menu_opt_var2 = StringVar()
menu_opt_var2.set(list_menu[0])
menu_opt2 = OptionMenu(left_framefruit, menu_opt_var2, *list_menu)
menu_opt2.configure(bg=color4,bd=1,relief=RIDGE , fg=color, font=("Times", 10))
menu_opt2.place(x=180, y=60)

menu_opt_var3 = StringVar()
menu_opt_var3.set(list_menu[0])
menu_opt3 = OptionMenu(left_framefruit, menu_opt_var3, *list_menu)
menu_opt3.configure(bg=color4,bd=1,relief=RIDGE , fg=color,  font=("Times", 10))
menu_opt3.place(x=180, y=100)

menu_opt_var4 = StringVar()
menu_opt_var4.set(list_menu[0])
menu_opt4 = OptionMenu(left_framefruit, menu_opt_var4, *list_menu)
menu_opt4.configure(bg=color4,bd=1,relief=RIDGE , fg=color,  font=("Times", 10))
menu_opt4.place(x=180, y=140)

menu_opt_var5 = StringVar()
menu_opt_var5.set(list_menu[0])
menu_opt5 = OptionMenu(left_framefruit, menu_opt_var5, *list_menu)
menu_opt5.configure(bg=color4,bd=1,relief=RIDGE , fg=color, font=("Times", 10))
menu_opt5.place(x=180, y=180)

menu_opt_var6 = StringVar()
menu_opt_var6.set(list_menu[0])
menu_opt6 = OptionMenu(left_framefruit, menu_opt_var6, *list_menu)
menu_opt6.configure(bg=color4,bd=1,relief=RIDGE , fg=color, font=("Times", 10))
menu_opt6.place(x=180, y=220)

menu_opt_var7 = StringVar()
menu_opt_var7.set(list_menu[0])
menu_opt7 = OptionMenu(left_framefruit, menu_opt_var7, *list_menu)
menu_opt7.configure(bg=color4,bd=1,relief=RIDGE , fg=color, font=("Times", 10))
menu_opt7.place(x=180, y=260)

menu_opt_var8 = StringVar()
menu_opt_var8.set(list_menu[0])
menu_opt8 = OptionMenu(left_framefruit, menu_opt_var8, *list_menu)
menu_opt8.configure(bg=color4,bd=1,relief=RIDGE , fg=color, font=("Times", 10))
menu_opt8.place(x=180, y=300)


# ================================BUTTONS FOR COMPUTATION and entries for pricing other items.=====================
entry_price_other_fruit_var = StringVar()
entry_price_other_fruit = Entry(root, textvariable=entry_price_other_fruit_var, width=1,bg=color3,borderwidth=0)
entry_price_other_fruit.place(x=100, y=5)

entry_price_other_juice_var = StringVar()
entry_price_other_juice = Entry(root, textvariable=entry_price_other_juice_var, width=1,bg=color3,borderwidth=0)
entry_price_other_juice.place(x=350, y=5)

total_btn_btm_frame = Button(root, text="TOTAL", width=10, height=2,bg=color4,bd=7,relief=RAISED,font=("Candra", 8,"bold") ,fg=color, command=lambda: total_sales())
total_btn_btm_frame.place(x=70, y=410)
receipt_btn_btm_frame = Button(root, text="Get Receipt", width=10,height=2, bg=color4,bd=7,relief=RAISED,font=("Candra", 8,"bold") ,fg=color,command=lambda: print_receipt())
receipt_btn_btm_frame.place(x=670, y=350)

book_k_btn_btm_frame = Button(root, text="Specific Product Deletion", width=20,height=2,bg=color4,bd=7,relief=RAISED,font=("Candra", 8,"bold") ,fg=color, command=lambda:  product_deletion())
book_k_btn_btm_frame.place(x=270, y=410)

camera_k_btn_btm_frame = Button(root, text="Camera Scanning", width=20,height=2,bg=color4,bd=7,relief=RAISED,font=("Candra", 8,"bold") ,fg=color, command=lambda:camera_scanning())
camera_k_btn_btm_frame.place(x=270, y=350)

rfid_k_btn_btm_frame = Button(root, text="Rfid Reader Scanning", width=20,height=2,bg=color4,bd=7,relief=RAISED,font=("Candra", 8,"bold") ,fg=color, command=lambda: rfid_read())
rfid_k_btn_btm_frame.place(x=460, y=350)

report_btn_btm_frame = Button(root, text="Proceed to checkout", width=20,height=2, bg=color4,bd=7,relief=RAISED,font=("Candra", 8,"bold") ,fg=color, state=DISABLED, command=lambda: checkout())
report_btn_btm_frame.place(x=460, y=410)

rst=PhotoImage(file="4back.png")
reset_btn_btm_frame = Button(root, image = rst , width=50,height=50, bg=color3,borderwidth=0, command=lambda: reset_sales())
reset_btn_btm_frame.place(x=670, y=410)

exit = PhotoImage(file = "cls.png")
exit_btn_btm_frame = Button(root, image = exit , width=50,height=50, bg=color3,borderwidth=0, command=lambda:exit_program())
exit_btn_btm_frame.place(x=740, y=410)

# ============================TITLE on top FRAME=======================
label_show = Label(root, text="Autonoumous Trolley With Smart Billing", bd=7, relief=RIDGE, width=35, height=2,
                   font=("Times", 12, "bold", "italic"), bg=color4, fg=color).place(x=240, y=0)

label_total_Purchase = Label(root, text=" Your Purchase ", bg=color3,bd=3,relief=RIDGE, font=("Times", 15, "bold"), fg=color)
label_total_Purchase.place(x=40, y=5)

root.mainloop()
